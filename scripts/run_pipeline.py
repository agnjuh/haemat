# run_pipeline.py — PBMC scRNA-seq ML pipeline (curated markers + leak-free split + gene-level + uncertainty)
import os, warnings, random, json, time
warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import scanpy as sc
import matplotlib.pyplot as plt
import scipy.sparse as sp

from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from sklearn.linear_model import LogisticRegression
from sklearn.neighbors import KNeighborsClassifier
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import (
    f1_score, balanced_accuracy_score, classification_report, confusion_matrix,
    brier_score_loss
)
from sklearn.calibration import calibration_curve

# config
RANDOM_SEED = 42
np.random.seed(RANDOM_SEED); random.seed(RANDOM_SEED)

DATA_PATH = "data/5k_pbmc_10x.h5ad"
CURATED_CSV = "markers_curated.csv"   # <- curated file
os.makedirs("results", exist_ok=True)

# helpers
def _to_dense(x):
    """Convert sparse/dense matrix to a dense numpy array."""
    return x.toarray() if sp.issparse(x) else np.asarray(x)

def get_counts_layer(a):
    """Return count matrix if available, else raw.X or X."""
    if "counts" in a.layers:
        return a.layers["counts"]
    if a.raw is not None:
        return a.raw.X
    return a.X

def _keep_gene(g):
    """Filter out overly generic/technical signals (HLA-, ribosomal, mitochondrial)."""
    return not (g.startswith(("HLA-", "RPL", "RPS", "MT-")))

def save_confusion(y_true, y_pred, labels, title, out_png, out_txt):
    """Save confusion matrix with a colorbar + text report."""
    # metrics to txt
    f1 = f1_score(y_true, y_pred, average="macro")
    bal = balanced_accuracy_score(y_true, y_pred)
    rpt = classification_report(y_true, y_pred, digits=3)
    with open(out_txt, "w") as f:
        f.write(rpt + f"\nmacro-F1={f1:.3f}, balanced-acc={bal:.3f}\n")

    # matrix
    cm = confusion_matrix(y_true, y_pred, labels=labels)

    # figures colourbars
    fig, ax = plt.subplots(figsize=(7.2, 5.8))
    im = ax.imshow(cm, aspect="auto")  # default cmap
    cb = fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
    cb.set_label("Count")

    ax.set_title(title)
    ax.set_xlabel("Predicted")
    ax.set_ylabel("True")
    ax.set_xticks(range(len(labels)))
    ax.set_yticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=45, ha="right")
    ax.set_yticklabels(labels)

    # annotate counts
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, str(cm[i, j]), ha="center", va="center")

    fig.tight_layout()
    fig.savefig(out_png, dpi=300)
    plt.close(fig)

    return {"macro_f1": f1, "balanced_acc": bal}

# curated marker loader
def load_curated_csv(path=CURATED_CSV, tiers=("core", "extended")):
    """
    Load curated markers from CSV → dict {cell_type: [GENE,...]}.

    Expected columns: cell_type,gene,tier,polarity,source,notes
    - Uses only rows where polarity is 'positive' and tier is in `tiers`.
    - Upper-cases gene symbols and de-duplicates per cell type.
    Also writes a flat table + simple provenance to results/.
    """
    assert os.path.exists(path), f"Curated marker file not found: {path}"
    df = pd.read_csv(path)
    # normalize headers
    df.columns = [c.strip().lower() for c in df.columns]
    req = {"cell_type","gene","tier","polarity"}
    missing = req - set(df.columns)
    assert not missing, f"Curated CSV missing columns: {sorted(missing)}"

    df = df[df["tier"].isin(tiers)].copy()
    df["polarity"] = df["polarity"].fillna("positive").str.lower()
    df = df[df["polarity"] == "positive"].copy()
    df["gene"] = df["gene"].astype(str).str.upper().str.strip()

    marker_sets = {
        ct: sorted(df.loc[df["cell_type"] == ct, "gene"].dropna().unique().tolist())
        for ct in df["cell_type"].unique()
    }

    # save provenance + flat export
    df.to_csv("results/markers_table.csv", index=False)
    prov = {
        "source": "curated CSV",
        "path": os.path.abspath(path),
        "tiers_used": list(tiers),
        "cell_types": sorted(marker_sets.keys()),
        "n_rows": int(len(df)),
    }
    with open("results/markers_provenance.json", "w") as f:
        json.dump(prov, f, indent=2)
    print("[markers] Using marker source → curated CSV")
    print("[markers] Saved → results/markers_provenance.json, results/markers_table.csv")
    return marker_sets

# load data
assert os.path.exists(DATA_PATH), f"Data not found: {DATA_PATH}"
adata = sc.read_h5ad(DATA_PATH)
adata.var_names_make_unique()
print(adata)

# labels (existing or proxy)
label_col = None
for c in adata.obs.columns:
    cl = c.lower()
    if ("cell" in cl and "type" in cl) or ("annotation" in cl):
        label_col = c
        break

# pull markers from curated CSV (no network)
marker_sets = load_curated_csv(CURATED_CSV, tiers=("core","extended"))

if label_col is None:
    print("No cell-type labels found. Building proxy labels via Leiden + marker scoring...")
    tmp = adata.copy()
    tmp.var_names_make_unique()
    sc.pp.normalize_total(tmp, target_sum=1e4); sc.pp.log1p(tmp)

    try:
        # HVG + PCA + neighbors + Leiden
        sc.pp.highly_variable_genes(tmp, n_top_genes=3000, flavor="seurat_v3")
        tmp_hvg = tmp[:, tmp.var.highly_variable].copy()
        sc.pp.scale(tmp_hvg, max_value=10)
        sc.tl.pca(tmp_hvg, n_comps=50, svd_solver="arpack")
        sc.pp.neighbors(tmp_hvg, n_neighbors=15, n_pcs=30)
        sc.tl.leiden(tmp_hvg, resolution=0.6, key_added="leiden")

        # marker scoring on the HVG object
        for ct, genes in marker_sets.items():
            present = [g for g in genes if g in tmp_hvg.var_names]
            if present:
                sc.tl.score_genes(tmp_hvg, present, score_name=f"score_{ct}", use_raw=False)
            else:
                tmp_hvg.obs[f"score_{ct}"] = 0.0

        # cluster → cell type map (based on the maximum of average marker scores)
        cluster2ct = {}
        for clab in sorted(tmp_hvg.obs["leiden"].unique(), key=lambda x: int(x)):
            sub = tmp_hvg[tmp_hvg.obs["leiden"] == clab]
            means = {ct: float(sub.obs[f"score_{ct}"].mean()) for ct in marker_sets}
            cluster2ct[clab] = max(means, key=means.get)

        # index-safe assignment
        adata.obs["cell_type"] = pd.Series(
            [cluster2ct[cl_] for cl_ in tmp_hvg.obs["leiden"]],
            index=tmp_hvg.obs.index,
            dtype="object"
        )

        # remaining (NaN) cells via marker argmax over the full tmp
        unlabeled_mask = adata.obs["cell_type"].isna()
        if unlabeled_mask.any():
            for ct, genes in marker_sets.items():
                present = [g for g in genes if g in tmp.var_names]
                if present:
                    sc.tl.score_genes(tmp, present, score_name=f"score_{ct}", use_raw=False)
                else:
                    tmp.obs[f"score_{ct}"] = 0.0
            scores = tmp.obs[[f"score_{ct}" for ct in marker_sets]].to_numpy()
            best_idx = scores.argmax(axis=1)
            labels = list(marker_sets.keys())
            proxy = pd.Series([labels[i] for i in best_idx], index=tmp.obs.index, dtype="object")
            adata.obs.loc[unlabeled_mask, "cell_type"] = proxy.loc[unlabeled_mask].values

    except Exception as e:
        print(f"Leiden unavailable or failed ({e}). Falling back to marker-only argmax.")
        sc.pp.normalize_total(tmp, target_sum=1e4); sc.pp.log1p(tmp)
        for ct, genes in marker_sets.items():
            present = [g for g in genes if g in tmp.var_names]
            if present:
                sc.tl.score_genes(tmp, present, score_name=f"score_{ct}", use_raw=False)
            else:
                tmp.obs[f"score_{ct}"] = 0.0
        scores = tmp.obs[[f"score_{ct}" for ct in marker_sets]].to_numpy()
        best_idx = scores.argmax(axis=1)
        labels = list(marker_sets.keys())
        adata.obs["cell_type"] = [labels[i] for i in best_idx]

    label_col = "cell_type"
    del tmp

# save class distribution
y_all = adata.obs[label_col].astype(str).values
pd.Series(y_all).value_counts().to_csv("results/class_distribution.csv")
print("Saved class distribution → results/class_distribution.csv")

# leak-free split + preprocessing
idx = np.arange(adata.n_obs)
idx_tr, idx_te = train_test_split(idx, test_size=0.2, random_state=RANDOM_SEED, stratify=y_all)
ad_tr = adata[idx_tr].copy()
ad_te = adata[idx_te].copy()

# normalise + log separately (deterministic)
for a in (ad_tr, ad_te):
    a.X = get_counts_layer(a).copy()
    sc.pp.normalize_total(a, target_sum=1e4)
    sc.pp.log1p(a)

# HVG on train only (fallback to cell_ranger if needed)
try:
    sc.pp.highly_variable_genes(ad_tr, n_top_genes=3000, flavor="seurat_v3")
except Exception:
    sc.pp.highly_variable_genes(ad_tr, n_top_genes=3000, flavor="cell_ranger")

hvg = [g for g in ad_tr.var.index[ad_tr.var["highly_variable"]] if g in ad_te.var_names]
ad_tr = ad_tr[:, hvg].copy()
ad_te = ad_te[:, hvg].copy()
pd.Series(hvg).to_csv("results/hvg_genes_train_only.csv", index=False)

# scaling + PCA (fit/train only)
Xtr = _to_dense(ad_tr.X)
Xte = _to_dense(ad_te.X)

scaler = StandardScaler(with_mean=True, with_std=True)
Xtr_s = scaler.fit_transform(Xtr)
Xte_s = scaler.transform(Xte)

pca = PCA(n_components=50, random_state=RANDOM_SEED)
X_train = pca.fit_transform(Xtr_s)
X_test  = pca.transform(Xte_s)

y_train = ad_tr.obs[label_col].astype(str).values
y_test  = ad_te.obs[label_col].astype(str).values
classes = np.unique(y_train)

# models + metrics
results = {}

# kNN
knn = KNeighborsClassifier(n_neighbors=15).fit(X_train, y_train)
pred_knn = knn.predict(X_test)
results["kNN"] = save_confusion(
    y_test, pred_knn, classes,
    "kNN – Confusion matrix",
    "results/confusion_knn.png",
    "results/report_knn.txt",
)

# Logistic Regression (multinomial)
lr = LogisticRegression(max_iter=500, multi_class="multinomial", random_state=RANDOM_SEED).fit(X_train, y_train)
pred_lr = lr.predict(X_test)
results["LogReg"] = save_confusion(
    y_test, pred_lr, classes,
    "Logistic Regression – Confusion matrix",
    "results/confusion_logreg.png",
    "results/report_logreg.txt",
)

# Random Forest
rf = RandomForestClassifier(n_estimators=400, random_state=RANDOM_SEED, n_jobs=-1).fit(X_train, y_train)
pred_rf = rf.predict(X_test)
results["RandomForest"] = save_confusion(
    y_test, pred_rf, classes,
    "Random Forest – Confusion matrix",
    "results/confusion_rf.png",
    "results/report_rf.txt",
)

pd.DataFrame(results).to_csv("results/metrics_summary.csv")
print("Saved metrics and confusion matrices → results/")

# gene-level interpretability (PCA→gene back-projection)
loadings = pca.components_.T                 # (n_genes, n_pc)
genes_all = np.array(ad_tr.var_names)
class_names = lr.classes_
summary_rows = []

for ci, cname in enumerate(class_names):
    w = lr.coef_[ci]                         # PC weights for class ci
    g_imp = np.abs(loadings @ np.abs(w))     # aggregate gene contribution
    df = (pd.DataFrame({"gene": genes_all, "importance": g_imp})
            .sort_values("importance", ascending=False))
    # filter out technical genes here too
    df_filt = df[df["gene"].map(_keep_gene)]
    df_filt.head(50).to_csv(f"results/top_genes_{cname}.csv", index=False)
    summary_rows.append({"cell_type": cname, "top10_genes": ", ".join(df_filt.head(10)["gene"].tolist())})

pd.DataFrame(summary_rows).to_csv("results/top_genes_summary.csv", index=False)
print("Saved gene-level rankings (filtered) → results/top_genes_*.csv")

# uncertainty & calibration
proba = lr.predict_proba(X_test)
yhat  = pred_lr
maxp  = proba.max(axis=1)

def cov_acc_at(t):
    keep = maxp >= t
    if keep.sum() == 0:
        return 0.0, np.nan
    return keep.mean(), (y_test[keep] == yhat[keep]).mean()

ts = np.linspace(0.5, 0.95, 10)
cov, acc = zip(*[cov_acc_at(t) for t in ts])

plt.figure()
plt.plot(ts, cov, label="coverage")
plt.plot(ts, acc, label="accuracy")
plt.xlabel("threshold on max-prob"); plt.legend(); plt.title("Coverage vs Accuracy")
plt.tight_layout(); plt.savefig("results/coverage_accuracy.png", dpi=300); plt.close()

# calibration curve: use the most frequent predicted class
ci = np.argmax(np.bincount(np.argmax(proba, axis=1)))
y_bin = (y_test == lr.classes_[ci]).astype(int)
prob_true, prob_pred = calibration_curve(y_bin, proba[:, ci], n_bins=10, strategy="uniform")

plt.figure()
plt.plot(prob_pred, prob_true, marker="o")
plt.plot([0,1],[0,1],"--")
plt.xlabel("Predicted probability"); plt.ylabel("True frequency")
plt.title(f"Calibration – class: {lr.classes_[ci]}")
plt.tight_layout(); plt.savefig("results/calibration_curve.png", dpi=300); plt.close()

with open("results/calibration_stats.txt","w") as f:
    f.write(f"Chosen class: {lr.classes_[ci]}\n")
    f.write(f"Brier score: {brier_score_loss(y_bin, proba[:,ci]):.4f}\n")

print("Saved uncertainty & calibration outputs → results/")
# Save probability outputs and labels for downstream threshold analysis
np.save("results/proba.npy", proba)
np.save("results/y_test.npy", y_test)
np.save("results/classes.npy", lr.classes_)
print("Saved proba/y_test/classes → results/")

# sparse gene-level LR directly on HVGs (cleaner markers)
lr_gene = LogisticRegression(
    penalty="l1", solver="saga", C=0.5, max_iter=2000, multi_class="multinomial", random_state=RANDOM_SEED
)
lr_gene.fit(Xtr_s, y_train)

coefs = lr_gene.coef_  # (n_classes, n_genes)
genes_hvg = np.array(ad_tr.var_names)

mask = np.array([_keep_gene(g) for g in genes_hvg])
genes_f = genes_hvg[mask]
coefs_f = coefs[:, mask]

summary_rows2 = []
for ci, cname in enumerate(lr_gene.classes_):
    imp = np.abs(coefs_f[ci])
    top10 = genes_f[imp.argsort()[::-1][:10]]
    summary_rows2.append({"cell_type": cname, "top10_genes": ", ".join(top10)})
pd.DataFrame(summary_rows2).to_csv("results/top_genes_summary_geneLR.csv", index=False)
print("Saved filtered gene-level rankings → results/top_genes_summary_geneLR.csv")

# DE (Wilcoxon) on train for cross-check
ad_train_for_de = ad_tr.copy()
ad_train_for_de.obs[label_col] = ad_train_for_de.obs[label_col].astype("category")
sc.tl.rank_genes_groups(ad_train_for_de, groupby=label_col, method="wilcoxon")
# Example manual export if needed:
# sc.get.rank_genes_groups_df(ad_train_for_de, group='T').head(20).to_csv("results/DE_T_wilcoxon_top20.csv", index=False)

# automatic gene annotation via MyGene.info (CSV + XLSX)
# (Network optional; this block fails gracefully and skips on outage.)
import requests

IN_CSV  = "results/top_genes_summary_geneLR.csv"
OUT_CSV = "results/top_genes_annotated.csv"
OUT_XLS = "results/annotated_marker_table.xlsx"

API_URL = "https://mygene.info/v3/query"
FIELDS  = "symbol,name,summary,entrezgene,ensembl.gene,go.BP.term,go.MF.term,go.CC.term"
SPECIES = "human"

def _explode_toplist(csv_path):
    if not os.path.exists(csv_path):
        print(f"[annot] Skip: {csv_path} not found.")
        return None
    dfi = pd.read_csv(csv_path)
    rows = []
    for _, r in dfi.iterrows():
        ct = str(r["cell_type"])
        genes = [g.strip() for g in str(r["top10_genes"]).split(",") if g.strip()]
        for g in genes:
            rows.append({"cell_type": ct, "gene": g})
    if not rows:
        print("[annot] Skip: empty gene list.")
        return None
    return pd.DataFrame(rows).drop_duplicates().reset_index(drop=True)

def _query_mygene_symbol(symbol: str, retries=3, pause=0.25, timeout=10):
    params = {"q": f"symbol:{symbol}", "species": SPECIES, "fields": FIELDS, "size": 1}
    for _ in range(retries):
        try:
            r = requests.get(API_URL, params=params, timeout=timeout)
            if r.status_code == 200:
                js = r.json()
                hits = js.get("hits", [])
                if hits:
                    return hits[0]
                return {}
        except Exception:
            pass
        time.sleep(pause)
    return {}

def _join_terms(obj):
    if isinstance(obj, list):
        terms = []
        for it in obj:
            if isinstance(it, dict) and "term" in it:
                terms.append(it["term"])
            elif isinstance(it, str):
                terms.append(it)
        return "; ".join(terms[:5])
    return ""

def _short_anno(row, width=90):
    text = (row.get("summary") or "").strip()
    if not text:
        text = (row.get("go_BP") or "").strip()
    if not text:
        text = (row.get("name") or "").strip()
    if not text:
        text = "—"
    text = " ".join(text.split())
    return (text[:width-1] + "…") if len(text) > width else text

def _write_xlsx(df_annot):
    try:
        import openpyxl
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        cols = ["cell_type","gene","annotation_short"]
        df_small = df_annot[cols].groupby("cell_type").head(10).reset_index(drop=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Annotated markers"

        header_fill = PatternFill("solid", fgColor="BDBDBD")
        bold_font   = Font(bold=True)
        left        = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ct_colors   = {"B":"C6DBEF","DC":"C7E9C0","Mono":"FDD0A2","NK":"F2B6C6","T":"D9C2F0"}

        # header
        for j, name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=j, value=name)
            cell.fill = header_fill; cell.font = bold_font; cell.alignment = left

        # rows
        for i, row in df_small.iterrows():
            r = i + 2
            ct = str(row["cell_type"])
            ws.cell(row=r, column=1, value=ct).fill = PatternFill("solid", fgColor=ct_colors.get(ct,"FFFFFF"))
            ws.cell(row=r, column=1).alignment = left
            ws.cell(row=r, column=2, value=str(row["gene"])).alignment = left
            ws.cell(row=r, column=3, value=str(row["annotation_short"])).alignment = left

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 90

        # full sheet
        ws2 = wb.create_sheet("Full_annotation")
        full_cols = ["cell_type","gene","symbol","name","summary","go_BP","go_MF","go_CC"]
        present = [c for c in full_cols if c in df_annot.columns]
        # header
        for j, name in enumerate(present, start=1):
            cell = ws2.cell(row=1, column=j, value=name)
            cell.fill = header_fill; cell.font = bold_font; cell.alignment = left
        # rows
        for i, row in df_annot[present].iterrows():
            r = i + 2
            for j, col in enumerate(present, start=1):
                ws2.cell(row=r, column=j, value=str(row[col]) if pd.notna(row[col]) else "")
                ws2.cell(row=r, column=j).alignment = left
        if "cell_type" in present: ws2.column_dimensions["A"].width = 12
        if len(present) >= 2: ws2.column_dimensions["B"].width = 18
        if "summary" in present:
            idx = present.index("summary") + 1
            ws2.column_dimensions[chr(64+idx)].width = 90

        wb.save(OUT_XLS)
        print(f"Saved (editable) → {OUT_XLS}")
    except Exception as e:
        print(f"[xlsx] WARNING: failed to write XLSX ({e})")

def build_annotated_outputs():
    df0 = _explode_toplist(IN_CSV)
    if df0 is None:
        return False

    cache_path = "results/_mygene_cache.json"
    cache = {}
    if os.path.exists(cache_path):
        try:
            cache = json.load(open(cache_path, "r"))
        except Exception:
            cache = {}

    uniq = sorted(df0["gene"].unique().tolist())
    recs = []
    for g in uniq:
        hit = cache.get(g)
        if hit is None:
            hit = _query_mygene_symbol(g)
            cache[g] = hit
            time.sleep(0.15)  # be kind to the API

        symbol  = hit.get("symbol", g)
        name    = hit.get("name", "")
        summary = hit.get("summary", "")
        entrez  = hit.get("entrezgene", "")

        ens = hit.get("ensembl", {})
        if isinstance(ens, dict):
            ens_gene = ens.get("gene", "")
        elif isinstance(ens, list) and len(ens) and isinstance(ens[0], dict):
            ens_gene = ens[0].get("gene", "")
        else:
            ens_gene = ""

        go = hit.get("go", {}) if isinstance(hit.get("go"), dict) else {}
        go_bp = _join_terms(go.get("BP", []))
        go_mf = _join_terms(go.get("MF", []))
        go_cc = _join_terms(go.get("CC", []))

        recs.append({
            "gene": g, "symbol": symbol, "name": name, "summary": summary,
            "entrezgene": entrez, "ensembl_gene": ens_gene,
            "go_BP": go_bp, "go_MF": go_mf, "go_CC": go_cc,
        })

    try:
        json.dump(cache, open(cache_path, "w"))
    except Exception:
        pass

    ann = pd.DataFrame(recs)
    df_annot = df0.merge(ann, on="gene", how="left")
    df_annot["annotation_short"] = df_annot.apply(_short_anno, axis=1)
    df_annot.to_csv(OUT_CSV, index=False)
    print(f"Saved CSV → {OUT_CSV}")

    _write_xlsx(df_annot)
    return True

try:
    ok = build_annotated_outputs()
    if not ok:
        print("[annot] Skipped: no gene list found to annotate.")
except Exception as e:
    print(f"[annot] WARNING: automatic annotation failed ({e})")

# minimalist top-gene outputs (symbols only)
# - Per-cell-type top-5 barplots
# - Union top-5 heatmap (fixed order) + CSV export

os.makedirs("results/plots", exist_ok=True)

# build importance table once from gene-level L1 logistic regression
imp_rows = []
try:
    for ci, cname in enumerate(lr_gene.classes_):
        imp = np.abs(coefs_f[ci])          # (n_genes_filtered,)
        order = np.argsort(imp)[::-1]
        for j in order[:100]:
            imp_rows.append({
                "cell_type": cname,
                "gene": str(genes_f[j]),
                "importance": float(imp[j]),
            })
    imp_df = pd.DataFrame(imp_rows)
except Exception as e:
    print(f"[plots] WARNING: could not build importance table ({e})")
    imp_df = pd.DataFrame(columns=["cell_type", "gene", "importance"])

if not imp_df.empty:
    # per-cell-type Top-5 barplots
    for ct in sorted(imp_df["cell_type"].unique()):
        sub = (imp_df[imp_df["cell_type"] == ct]
               .sort_values("importance", ascending=False)
               .drop_duplicates("gene")
               .head(5)
               .copy())
        if sub.empty:
            continue
        plt.figure(figsize=(6.5, 3.6))
        plt.barh(range(len(sub)), sub["importance"].values)
        plt.gca().invert_yaxis()
        plt.yticks(range(len(sub)), sub["gene"].tolist(), fontsize=10)
        plt.xlabel("Importance (|L1 coefficients|)")
        plt.title(ct)
        plt.tight_layout()
        out_png = f"results/plots/top5_{ct}.png"
        plt.savefig(out_png, dpi=300)
        plt.close()
        print(f"Saved → {out_png}")

    # union Top-5 heatmap + CSV (fixed order)
    fixed_ct_order = ["B", "DC", "Mono", "NK", "T"]
    cts_present = [ct for ct in fixed_ct_order if ct in imp_df["cell_type"].unique()]
    if not cts_present:
        cts_present = sorted(imp_df["cell_type"].unique())

    top_per_ct = {}
    for ct in cts_present:
        genes_top5 = (imp_df[imp_df["cell_type"] == ct]
                      .sort_values("importance", ascending=False)
                      .drop_duplicates("gene")
                      .head(5)["gene"].tolist())
        top_per_ct[ct] = genes_top5

    all_genes = []
    for ct in cts_present:
        for g in top_per_ct[ct]:
            if g not in all_genes:
                all_genes.append(g)

    M = np.zeros((len(all_genes), len(cts_present)), dtype=float)
    for j, ct in enumerate(cts_present):
        sub = imp_df[imp_df["cell_type"] == ct].set_index("gene")["importance"]
        for i, g in enumerate(all_genes):
            M[i, j] = float(sub.get(g, 0.0))

    # CSV export
    hm_df = pd.DataFrame(M, index=all_genes, columns=cts_present)
    hm_csv = "results/top5_union_heatmap.csv"
    hm_df.to_csv(hm_csv)
    print(f"Saved → {hm_csv}")

    # heatmap (PNG + PDF) + colourbar
    plt.figure(figsize=(1.1*len(cts_present) + 4, 0.45*len(all_genes) + 2))
    im = plt.imshow(M, aspect="auto")
    cb = plt.colorbar(im, fraction=0.025, pad=0.02)
    cb.set_label("Importance (|L1|)")
    plt.xticks(range(len(cts_present)), cts_present)
    plt.yticks(range(len(all_genes)), all_genes, fontsize=8)
    plt.title("Top-5 union per cell type — importance heatmap")
    plt.tight_layout()
    out_png = "results/plots/top5_union_heatmap.png"
    out_pdf = "results/plots/top5_union_heatmap.pdf"
    plt.savefig(out_png, dpi=300)
    plt.savefig(out_pdf)
    plt.close()
    print(f"Saved → {out_png}, {out_pdf}")
else:
    print("[plots] Skipped: no importance data available.")

# --- FINAL LINE ---
print("DONE.")